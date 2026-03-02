"""
AR Aging Excel Generator
Queries FO_PRODUCTION_DB gold models via Snowflake Python connector
and produces a two-tab formatted workbook:
  Tab 1 — AR Aging Summary (owner-level aggregation)
  Tab 2 — AR Aging Detail (invoice-level)
"""

import snowflake.connector
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import tomllib, pathlib, sys, re
from datetime import date

# ── Connection ────────────────────────────────────────────────────────────────
config_path = pathlib.Path.home() / ".snowflake" / "config.toml"
with open(config_path, "rb") as f:
    config = tomllib.load(f)

conn_cfg = config["connections"]["formentera"]

conn = snowflake.connector.connect(
    account=conn_cfg["account"],
    user=conn_cfg["user"],
    private_key_file=conn_cfg.get("private_key_file"),
    role=conn_cfg.get("role", "DBT_ROLE"),
    warehouse=conn_cfg.get("warehouse", "TRANSFORMING"),
    database="FO_PRODUCTION_DB",
    schema="GOLD_FINANCIAL",
)

cur = conn.cursor()

# ── Parameters ────────────────────────────────────────────────────────────────
CUTOFF_DATE = "2026-02-26"
COMPANY     = "200"
OUT_FILE    = f"c:/Users/MichaelShiraz(CONTRA/Downloads/AR Aging - Pre {CUTOFF_DATE}.xlsx"

# ── Styles ────────────────────────────────────────────────────────────────────
DARK_BLUE   = "1F4E79"
LIGHT_BLUE  = "D9E1F2"
GRAY        = "595959"
CURRENCY    = '$#,##0.00'
RIGHT       = Alignment(horizontal='right')
CENTER      = Alignment(horizontal='center')

def header_style(cell, bg=DARK_BLUE):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = CENTER

def subtotal_style(cell, col_idx, currency_cols):
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    if col_idx in currency_cols:
        cell.number_format = CURRENCY
        cell.alignment = RIGHT

def grand_total_style(cell, col_idx, currency_cols):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
    if col_idx in currency_cols:
        cell.number_format = CURRENCY
        cell.alignment = RIGHT

def write_report_header(ws, title, subtitle, col_count):
    col = get_column_letter(col_count)
    ws.merge_cells(f"A1:{col}1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = CENTER
    ws.merge_cells(f"A2:{col}2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(italic=True, size=10, color=GRAY)
    ws["A2"].alignment = CENTER
    ws.append([])

# ── Tab 1: AR Aging Summary (owner-level) ─────────────────────────────────────
print("Querying AR Aging Summary...")
cur.execute(f"""
    SELECT
        company_code,
        owner_code,
        owner_name,
        COUNT(*)                                        AS invoice_count,
        SUM(total_invoice_amount)                       AS total_billed,
        SUM(balance_90_plus_days)                       AS over_90,
        SUM(balance_61_90_days)                         AS days_61_90,
        SUM(balance_31_60_days)                         AS days_31_60,
        SUM(current_balance + balance_1_30_days)        AS days_current_30,
        SUM(remaining_balance)                          AS total_outstanding
    FROM GOLD_FCT_AR_AGING_SUMMARY
    WHERE include_record = 1
      AND company_code = '{COMPANY}'
      AND invoice_date < '{CUTOFF_DATE}'
    GROUP BY 1, 2, 3
    HAVING SUM(remaining_balance) <> 0
    ORDER BY owner_name
""")
summary_rows = cur.fetchall()
print(f"  {len(summary_rows)} owners with outstanding balance")

# ── Tab 2: AR Aging Detail (invoice-level) ────────────────────────────────────
print("Querying AR Aging Detail...")
cur.execute(f"""
    SELECT
        company_code,
        owner_code,
        owner_name,
        invoice_number,
        invoice_date,
        invoice_type,
        CASE WHEN hold_billing THEN 'Yes' ELSE '' END   AS hold_billing,
        total_invoice_amount,
        balance_90_plus_days,
        balance_61_90_days,
        balance_31_60_days,
        current_balance + balance_1_30_days             AS days_current_30,
        remaining_balance
    FROM GOLD_FCT_AR_AGING_SUMMARY
    WHERE include_record = 1
      AND company_code = '{COMPANY}'
      AND invoice_date < '{CUTOFF_DATE}'
      AND remaining_balance <> 0
    ORDER BY owner_name, invoice_date, invoice_number
""")
detail_rows = cur.fetchall()
print(f"  {len(detail_rows)} invoices with outstanding balance")
cur.close()
conn.close()

# ── Build Workbook ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ── TAB 1: Summary ─────────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "AR Aging Summary"

write_report_header(
    ws1,
    "Formentera Operations LLC — AR Aging Summary by Owner",
    f"Company {COMPANY}  |  Invoice Date < {CUTOFF_DATE}  |  Owners with outstanding balance only",
    10
)

headers = ["Company", "Owner Code", "Owner Name", "Invoices",
           "Total Billed", "Over 90 Days", "61-90 Days",
           "31-60 Days", "30 Days-Current", "Total Outstanding"]
ws1.append(headers)
h_row = ws1.max_row
for c in range(1, 11):
    header_style(ws1.cell(h_row, c))

currency_cols_s = {5, 6, 7, 8, 9, 10}
grand = [0] * 10

for row in summary_rows:
    vals = list(row)
    ws1.append(vals)
    r = ws1.max_row
    for c, v in enumerate(vals, 1):
        cell = ws1.cell(r, c)
        if c in currency_cols_s:
            cell.number_format = CURRENCY
            cell.alignment = RIGHT
        if c == 4:
            cell.alignment = CENTER
    for i, c in enumerate(sorted(currency_cols_s)):
        grand[c - 1] += float(vals[c - 1] or 0)

# Grand total
ws1.append([])
gt_vals = ["", "", "GRAND TOTAL", sum(r[3] for r in summary_rows)] + \
          [grand[c - 1] for c in sorted(currency_cols_s)]
ws1.append(gt_vals)
r = ws1.max_row
for c in range(1, 11):
    grand_total_style(ws1.cell(r, c), c, currency_cols_s)
ws1.cell(r, 4).alignment = CENTER

ws1.freeze_panes = "A5"
for i, w in enumerate([10, 12, 38, 10, 16, 14, 12, 12, 16, 16], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ── TAB 2: Detail ─────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("AR Aging Detail")

write_report_header(
    ws2,
    "Formentera Operations LLC — AR Aging Detail by Invoice",
    f"Company {COMPANY}  |  Invoice Date < {CUTOFF_DATE}  |  Outstanding invoices only",
    13
)

headers2 = ["Company", "Owner Code", "Owner Name", "Invoice #",
            "Invoice Date", "Invoice Type", "Hold Billing",
            "Invoice Amount", "Over 90 Days", "61-90 Days",
            "31-60 Days", "30 Days-Current", "Total Outstanding"]
ws2.append(headers2)
h_row2 = ws2.max_row
for c in range(1, 14):
    header_style(ws2.cell(h_row2, c))

currency_cols_d = {8, 9, 10, 11, 12, 13}
prev_owner, owner_totals = None, {c: 0.0 for c in currency_cols_d}
prev_owner_name = ""

for row in detail_rows:
    vals = list(row)
    owner = vals[1]

    if prev_owner is not None and owner != prev_owner:
        # Owner subtotal
        st = ["", prev_owner, f"  SUBTOTAL — {prev_owner_name}", "", "", "", ""] + \
             [owner_totals[c] for c in sorted(currency_cols_d)]
        ws2.append(st)
        sr = ws2.max_row
        for c in range(1, 14):
            subtotal_style(ws2.cell(sr, c), c, currency_cols_d)
        owner_totals = {c: 0.0 for c in currency_cols_d}

    prev_owner = owner
    prev_owner_name = vals[2]

    ws2.append(vals)
    r = ws2.max_row
    for c in range(1, 14):
        cell = ws2.cell(r, c)
        if c in currency_cols_d:
            cell.number_format = CURRENCY
            cell.alignment = RIGHT
            owner_totals[c] += float(vals[c - 1] or 0)

# Last owner subtotal
if prev_owner:
    st = ["", prev_owner, f"  SUBTOTAL — {prev_owner_name}", "", "", "", ""] + \
         [owner_totals[c] for c in sorted(currency_cols_d)]
    ws2.append(st)
    sr = ws2.max_row
    for c in range(1, 14):
        subtotal_style(ws2.cell(sr, c), c, currency_cols_d)

# Grand total
ws2.append([])
gt2 = ["", "", "GRAND TOTAL", "", "", "", ""] + \
      [sum(float(r[c - 1] or 0) for r in detail_rows) for c in sorted(currency_cols_d)]
ws2.append(gt2)
r = ws2.max_row
for c in range(1, 14):
    grand_total_style(ws2.cell(r, c), c, currency_cols_d)

ws2.freeze_panes = "A5"
for i, w in enumerate([10, 12, 35, 14, 14, 14, 12, 16, 14, 12, 12, 16, 16], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(OUT_FILE)
print(f"\nSaved: {OUT_FILE}")
print(f"  Summary tab: {len(summary_rows)} owners")
print(f"  Detail tab:  {len(detail_rows)} invoices")
